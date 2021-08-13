from PySimpleGUI.PySimpleGUI import R
from openpyxl import Workbook as xlwkbk
from openpyxl import load_workbook
import sys
from openpyxl_image_loader import SheetImageLoader
try: # if using functions from main.py
    from main_build.images import images
    from main_build.dependencies import helpers
    from main_build.dependencies import filestream
    from main_build.dependencies.data import *
except: # if using functions from this file 
    sys.path.insert(0,'../')
    from images import images 
    import helpers 
    import filestream

import PySimpleGUI as sg

class VersionInfo():
    def __init__(self):
        pass 
    def need_to_update(self)->bool:
        """
        Either creates a .metadata file or reads from it to get the current version
        Writing will happen when the user firsts opens the application
            -> Will get the information from the github api and find the most recent version
            -> Serializes it to .metadata and hides the file from plain sight 
        Reading will happen when the user opens the application any other time 
            -> Will query the github api to check version with .metadata
        """
        import os
        import requests
        latest_version_num = 0 
        try:
            # this will happen everytime the user calles the function (that is not the first time) 
            metadata_f = open('.metadata', 'r')
            cur_version = float(metadata_f.read()) 
            try: # get the latest version 
                response = requests.get("https://api.github.com/repos/panariellop/mathcad_auto/releases")
                latest_version_num = float(response.json()[0]['tag_name'])
                print(cur_version, latest_version_num) 
            except: pass  # in case of no internet 
            if latest_version_num > cur_version:
                return True 
            else:
                return False 
        except:
            # this will happen on the first time the application is opened by the user  
            try: 
                response = requests.get("https://api.github.com/repos/panariellop/mathcad_auto/releases")
                version = response.json()[0]['tag_name']
                metadata_f = open('.metadata', 'a')
                metadata_f.write(f'{version}')
                # hides the file from the windows file manager (invisible to user) 
                os.system('attrib +H *.metadata /S')
                return False
            except: # might happen if there is no internet connection 
                return False 

    def get_cur_version(self)->int:
        try:
            metadata_f = open('.metadata', 'r')
            cur_version = float(metadata_f.read())
            return cur_version
        except:
            #for some reason if this is called before self.need_to_update
            self.need_to_update()
            return self.get_cur_version()
            
class SelectTemplates():
    """
    This is the select templates window that users can use to edit or add
    templates to the program.
    """
    def __init__(self):
        try: # running from main.py 
            from main_build.dependencies.gui import VersionInfo
        except: pass # running locally 
        self.database = ""
        self.save_to_database = True
        self.excel = ""
        self.template_layout = []
        self.templates = {}
        self.wall_template = ""
        self.floor_template = ""
        self.wallfloor_template = ""
        self.ceiling_template = ""
        self.images = dict()
        self.output_folder = None
        self.can_continue = False
        self.version = VersionInfo() # instantiate 
        self.version = self.version.get_cur_version() # get actual value 

    def choose_templates(self):
        sg.theme("Reddit")
        window = sg.Window("Choose Template Files", [
            self.template_layout,
            [sg.Button("Continue", key = "continue", button_color = "green", enable_events = True)]
        ], icon = images.ma_logo_png)

        while True: #logic loop
            event, values = window.read()
            if event == 'OK' or event == sg.WIN_CLOSED:
                window.close()
                return False
            if event == "continue":
                #set each template to the corresponding GUI item
                errors = []
                for k, v in self.templates.items():
                    if helpers.check_file_type(values[k+"_name"], "mcdx"):
                        self.templates[k] = values[k+"_name"]
                    else:
                        errors.append(k+" mounting template must be a .mcdx file.")
                #error management
                if len(errors) < 1:
                    window.close()
                    return self.templates
                else:
                    popup = Popup("Error", " ".join(errors))
                    popup.alert()


    def display_and_update(self):
        """
        Displays the window and will update the object variables when the user hits continue
        """
        sg.theme("Reddit")
        window = sg.Window("Mathcad Automation", [
            [sg.Image(data = images.tt_logo)],
            [sg.Text(f'Version: {self.version}', size = (10,1))], 
            [sg.Frame("Choose Excel File*", [[sg.FileBrowse("1. Browse", key="excel_file", enable_events=True),
                                              sg.InputText(self.excel, key="excel_name", size=(30, 1),
                                                           background_color='white', enable_events=True),],

                                              [sg.Button("2. Choose Templates", key = "choose_templates", enable_events = True)],
                                            ])],
            [sg.Frame("Choose Database File (Optional)", [[sg.FileBrowse(key="database_file", enable_events=True),
                                                sg.InputText(self.database, key="database_name", size=(30, 1),
                                                             background_color='white', enable_events=True)], ])],
            [sg.Checkbox("Save to database?", key="save_to_database", default=self.save_to_database,
                         tooltip="If selected, details about the generated report will be saved to a database.")],

            [sg.Button("3. Continue", key="continue", button_color="green")],

        ], icon=images.ma_logo_png)
        """Listen for events"""
        while True: #logic loop
            event, values = window.read()
            if event == 'OK' or event == sg.WIN_CLOSED:
                window.close()
                return False
            else:
                if event == "choose_templates":
                    """
                    load the excel file and search for different mounting locations
                    """
                    if helpers.check_file_type(values['excel_name'], 'xlsx'):
                        #get the mounting locations
                        equipment = filestream.get_eqpt_from_xl(values['excel_name'])
                        if len(equipment.items) > 0:
                            self.excel = values['excel_name']
                            self.template_layout = []
                            for idx, item in enumerate(equipment.mounting_locations):
                                if item not in self.templates:
                                    self.templates[item] = "" # instanitate each templates to be empty string
                                #create the layout for the choose templates window
                                self.template_layout += [
                                    [sg.Text("Choose " + item + " mounting template:")],
                                    [sg.FileBrowse(key= item, enable_events=True),
                                      sg.InputText(self.templates[item], key=item + "_name",
                                      size=(60, 1), background_color='white', enable_events=True)],
                                ]
                            if self.choose_templates():
                                self.can_continue = True


                if event == "continue":  # user has input all information
                    # list of all the user errors when they input information
                    errors = list()
                    # validate the filepaths
                    self.save_to_database = values['save_to_database']
                    #check if the database is a .csv file - if not, append error list
                    if helpers.check_file_type(values['database_name'], 'csv') or values['database_name'] == "":
                        self.database = values['database_name']
                    else:
                        errors.append("Database file must be a .csv file.")
                    #check if excel file is .xlsx
                    if helpers.check_file_type(values['excel_name'], 'xlsx'):
                        self.excel = values['excel_name']
                    else:
                        errors.append("Excel file must be a .xlsx file.")
                    # Error handling
                    if len(errors) == 0 and self.can_continue: #no errors
                        window.close()
                        return True
                    elif self.can_continue == False:
                        alert = Popup("Error", "You must choose template files.")
                        alert.alert()
                        continue
                    else: #show errors in a popup window
                        alert = Popup("Errors", "\n".join(errors))
                        alert.alert()
                        continue

    def dev_get_xl_and_templates(self):
        """
        Skips the choose file menu by setting the excel file and template files to files already in system
        """
        self.excel = "C:/Users/Owner/Desktop/mathcad_auto/templates/dev_templates/starter_template_copy.xlsx"
        self.templates = {'Wall, Floor': 'C:/Users/Owner/Desktop/mathcad_auto/templates/dev_templates/starter_template_document_mathcad_7.mcdx',
                          'Ceiling': 'C:/Users/Owner/Desktop/mathcad_auto/templates/dev_templates/starter_template_document_mathcad_7.mcdx'
        }

    def get_images_from_xl(self, num_images: int):
        """
        Gets all the preview images from the excel file and saves a dictionary of
        mounting locations and binary images in self.images
        <num_images:int> Corresponds to the number of mounting locations,
        indicated in the preview_images portion of the input excel document
        """
        wb = load_workbook(self.excel)
        sheet = wb['preview_images']

        image_loader = SheetImageLoader(sheet)
        for idx, row in enumerate(sheet.iter_rows(values_only=True)): 
            try:
                image = image_loader.get('B' + str(idx+1) )
                import io 
                buf = io.BytesIO()
                image.save(buf, format = "PNG")
                self.images[row[0]] = buf.getvalue()
            except Exception as e: 
                print(e)

class Popup():
    """
    Simple class to create a popup window. Used to create a popup to input filenames,
    display errors, etc.
    """
    def __init__(self, title, message=""):
        self.title = title
        self.message = message

    def confirm(self) -> bool:
        """
        Confirmation window - if you want to confirm with the user yes/no
        """
        popup = sg.Window(self.title, [
            [sg.Text("")],
            [sg.Text(self.message), ],
            [sg.Text("")],
            [sg.Button("YES", key="YES", button_color='green'), sg.Button("NO", key="NO", button_color='red')],
        ],
                          icon=images.ma_logo_png, )
        while True:
            event, values = popup.read()
            if event == "OK" or event == "NO" or event == sg.WIN_CLOSED:
                popup.close()
                return False
            if event == "YES":
                popup.close()
                return True

    def alert(self):
        """
        Alert window - alert the user with a message
        """
        popup = sg.Window(self.title, [
            [sg.Text("")],
            [sg.Text(self.message)],
            [sg.Button("OK", key='OK', button_color='green')],
        ],
                          icon=images.ma_logo_png, )
        while True:
            event, values = popup.read()
            if event == "OK" or event == sg.WIN_CLOSED:
                popup.close()
                break
            else:
                pass

    def take_input(self, trailing_text: str) -> str:
        """
        Input window - if you want the user to input some information
        """
        popup = sg.Window(self.title, [
            [sg.Text("")],
            [sg.Text(self.message), sg.InputText(key="input", size=(15, 1)), sg.Text(trailing_text)],
            [sg.Button("OK", key='OK', button_color='green')],
        ],
                          icon=images.ma_logo_png, )
        while True:
            event, values = popup.read()
            if event == "OK" or event == sg.WIN_CLOSED:
                popup.close()
                return values['input']

    def take_filepath_input(self, output_folder)->str:
        popup = sg.Window(self.title, [
            [sg.Text(self.message)], 
            [sg.FolderBrowse("Browse", key="file", enable_events=True),
            sg.InputText(output_folder, key="filename", size=(45, 1),background_color='white', enable_events=True)
            ],
            [sg.Button("Continue", key="Continue")],
        ],
        icon=images.ma_logo_png, )
        
        while True:
            event, values = popup.read()
            if event == "OK" or event == "CANCEL" or event == sg.WIN_CLOSED:
                popup.close()
                return output_folder 
            if event == "Continue":
                popup.close()
                return values['filename']

    def image(self, image):
        """
        Creates a popup window with title and image
        """
        popup = sg.Window(self.title, [
            [sg.Image(data=image)]
        ])
        while True: #logic loop
            """
            Window will stay active until the user exits the window or they click the OK buttion
            """
            event, values = popup.read()
            if event == "OK" or event == sg.WIN_CLOSED:
                popup.close()
                break
    def link(self, description_text, link):
        """
        Creates a popup with a clickable link
        """
        popup = sg.Window(self.title, [
            [sg.Text("")],
            [sg.Multiline(self.message, s = (80,10), disabled = True, text_color = "black", background_color = "white")],
            [sg.Button(description_text, enable_events = True, key = "LINK")], 
            [sg.Text("")],
            [sg.Text("")],
        ], icon = images.ma_logo_png)
        while True:
            event, values = popup.read()
            if event == "OK" or event == sg.WIN_CLOSED:
                popup.close()
                break
            if event == "LINK":
                popup.close()
                import webbrowser
                webbrowser.open(link)
                break
class LoadingIndicator():
    def __init__(self, num_tasks):
        self.progress = 0
        self.num_tasks = num_tasks
        self.window = None 
    def render(self):
        self.window = sg.Window("Progress", [[
            sg.ProgressBar(self.num_tasks, orientation='h', size=(20, 20), key='progressbar')],
                                             [sg.Text(f'{self.progress}/{self.num_tasks} - Tasks Completed', key = "text", size = (20,1))],
                                             ]
                                , icon = images.ma_logo_png)
        event, values = self.window.read(timeout = 0.1)
        if event == 'Cancel'  or event == sg.WIN_CLOSED:
            self.window.close()
    def update(self):
        progress_bar = self.window['progressbar']
        self.window['text'].update(f'{self.progress+1}/{self.num_tasks} - Tasks Completed')
        self.progress +=1
        progress_bar.UpdateBar(self.progress)
    def close(self):
        if self.window != None:
            self.window.close()

class ViewReports():
    def __init__(self):
        self.window = None 
        self.database_file = None
        self.reports = [] 
        self.update_reports()
        self.render()
    def get_reports_attribute(self, attribute:str)->list():
        """
        Will generate a list of values given a key
        ex: File Names, Dates, etc
        """
        out = []
        for report in self.reports: #this is a list 
            for key, val in report.items(): #this is a dict
                if str(key) == attribute:
                    out.append(report[key])
        return out
    def get_reports_as_list(self, category = "", filter="", inclusive = True)->list():
        out = []
        filter = filter.upper()
        for report in self.reports:
            report_vals = []
            is_filtered = False 
            for key, val in report.items():
                report_vals.append(val)
                """
                Filter = wall
                val = wall,floor
                """
                if inclusive == True: 
                    if filter in val.upper() and category == key: #f = wall, v = wall,floor OK 
                        is_filtered = True 
                elif inclusive == False:
                    if filter == val.upper() and category == key: #f = wall, v = wall OK; K
                        is_filtered = True 
                if filter == "" or category == "":
                    is_filtered = True
            if len(report_vals) > 0 and is_filtered: out.append(report_vals)
        return out 

    def sort_by(self, reports:list, category:int, is_increasing = False)->list: 
        """
        Sorts the reports by category increasing or decreasing
        """
        if len(reports) == 0: return reports 
        sorted = [] 
        #create identifiers, put into multidimentsional array, then sort using arr.sort()
        multi = list()
        for idx, report in enumerate(reports):
            multi.append([report[category], idx]) 
        if is_increasing:
            multi.sort() 
        else:
            multi.sort(reverse=True)
        for m in multi:
            sorted.append(reports[m[1]])
        return sorted 

    def add_report(self, to_append):
        """
        Adds a report to the class
        """
        self.reports.append(to_append)

    def update_reports(self):
        """
        Check if there are new reports that we need to be aware of
        """
        self.reports = []
        csv_file = None 
        if self.database_file == None or self.database_file == "":
            cur_path = "./"
        else:
            cur_path = self.database_file
        import os 
        depth = 2 
        for root, dirs, files in helpers.walklevel(cur_path, depth = 2):
            for name in files:
                if name.endswith("csv"): # find the csv file 
                    csv_file = os.path.join(root, name)
        
        #parse the csv file 
        import csv
        if csv_file == None:
            csv_file = self.database_file
        try:
            with open(csv_file, "r", newline="") as f:
                reader = csv.DictReader(f)
                for line in reader:
                    self.add_report(line)
        except: #there is no file 
            pass 
        if csv_file: return True 
        else: return False 
    
    def render(self):
        """
        Render the GUI
        """
        sg.theme('Reddit')
        headings = []
        if len(self.reports) > 0 :
            for key, val in self.reports[0].items():
                headings.append(key)
        else: 
            status = self.choose_database_file()
            if not status: 
                return 
            else: 
                self.render()
        layout = [[
            [sg.Frame("Filter", [[sg.Combo(headings, default_value = headings[0], key= "search_scope", enable_events = True), 
            sg.InputText(key = "search_input", enable_events = True), 
            sg.Text("Inclusive?"), 
            sg.Combo(['Yes', 'No'], default_value = ['Yes'], key="is_inclusive", enable_events = True), 
            sg.Button("Clear", key="Clear")], ])],

            [sg.Frame("Sort By", [[
                sg.Combo(headings, default_value = headings[0], key= "sort_category", enable_events = True), 
                sg.Text("Increasing?"), 
                sg.Combo(['Yes', 'No'], default_value = ['Yes'], key="is_increasing", enable_events = True), 
            ]])], 
            
            [sg.Table(values = self.get_reports_as_list(), 
                headings = headings, 
                auto_size_columns = True, 
                max_col_width = 75, 
                def_col_width = 25, 
                num_rows = 45, 
                display_row_numbers = True, 
                enable_events = True, 
                key = "table",
                select_mode=['browse']
                )], 
        ]]

        menu = [['&File', ['&Select Database File']], ['&Export', ["Export Visible Data to Clipboard"]]]
        layout += [[sg.Menu(menu)]]       
        self.window = sg.Window("View Reports", layout, icon = images.ma_logo_png)

        while True:
            event, values = self.window.read()
            if event == 'Cancel'  or event == sg.WIN_CLOSED:
                self.window.close()
                break 

            if event == "search_scope" or "search_input" or "is_inclusive":
                if values['is_inclusive'] == "Yes":
                    is_inclusive = True
                else: is_inclusive = False 
                new_values = self.get_reports_as_list(values['search_scope'], values['search_input'], is_inclusive)
                category = headings.index(values['sort_category'])
                if values['is_increasing'] == "Yes":
                    sorted = self.sort_by(new_values, category, is_increasing=True)
                else:
                    sorted = self.sort_by(new_values, category, is_increasing=False)
                if len(new_values) > 0: 
                    self.window['table'].update(values = sorted)

            if event == "Clear": 
                values['search_input'] = ""
                self.window['search_input'].update("")
                self.window['is_inclusive'].update("Yes")
                self.window['table'].update(values = self.get_reports_as_list())

            if event == "Select Database File":
                self.choose_database_file()
                new_values = self.get_reports_as_list(values['search_scope'], values['search_input'], is_inclusive)
                if len(new_values) > 0: 
                    self.window['table'].update(values = new_values)

            if event == "Export Visible Data to Clipboard":
                import pandas 
                reports = self.get_reports_as_list(values['search_scope'], values['search_input'], is_inclusive)
                category = headings.index(values['sort_category'])
                if values['is_increasing'] == "Yes":
                    sorted = self.sort_by(reports, category, is_increasing=True)
                else:
                    sorted = self.sort_by(reports, category, is_increasing=False)
                df = pandas.DataFrame(data = sorted, columns = headings)
                df.to_clipboard()

            if event == "is_increasing" or event == "sort_category":
                reports = self.get_reports_as_list(values['search_scope'], values['search_input'], is_inclusive)
                category = headings.index(values['sort_category'])
                if values['is_increasing'] == "Yes":
                    sorted = self.sort_by(reports, category, is_increasing=True)
                else:
                    sorted = self.sort_by(reports, category, is_increasing=False)
                self.window['table'].update(values = sorted)
                continue 
    
    def choose_database_file(self):
        sg.theme("Reddit")
        window = sg.Window("Choose Database File", [
            [sg.FileBrowse("Browse", key="file", enable_events=True),
            sg.InputText(key="filename", size=(45, 1),background_color='white', enable_events=True)
            ],
            [sg.Button("Continue", key="Continue")],
            ]
                                            

        , icon = images.ma_logo_png)
        while True:
            event, values = window.read()
            if event == 'Cancel'  or event == sg.WIN_CLOSED:
                window.close()
                return False 
            if event == "Continue" and (values['filename'].endswith("csv") or values['filename'] == ""):
                self.database_file = values['filename']
                if self.update_reports():
                    window.close()
                    return True 


if __name__ == "__main__":
    view = ViewReports()
