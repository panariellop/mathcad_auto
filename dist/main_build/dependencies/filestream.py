"""
All the functions that pull information from or save information to files 
"""
try: # when running from main.py 
    from main_build.dependencies.data import Equipment
    from main_build.dependencies.gui import Popup 
    from main_build.dependencies.filestream import DebugLogger
except: # when runnning from current directory 
    #from data import Equipment
    #from gui import Popup 
    pass
from openpyxl import Workbook as xlwkbk
from openpyxl import load_workbook
from datetime import date
import csv

def save_eqpt_to_csv(values, filepath, unique_report_name):
    """
    Saves the equipment currently being reported to a csv file
    """
    debug = DebugLogger()
    cur_date = date.today()
    header = False
    # check if header needs to be generated (happens when the file hasn't been
    # created before)
    try:
        with open(filepath, "r", newline="") as f:
            pass
    except Exception as e:
        header = True #header needs to be generated
    with open(filepath, "a", newline="") as f:
        csv_writer = csv.writer(f)
        #write the header line
        if header: csv_writer.writerow(["Date", "Project Number", "Equipment Number", "Equipment Tags", "Tags", "Name", "Mounting Location", "File Name"])
        try:
            new_row = [ #create the new row
                cur_date,
                values['project_number'][0],
                values['eqpt_number'][0],
                values['eqpt_tags'][0].upper(), 
                values['tags'][0].upper(),
                values['eqpt_name'][0].upper(),
                values['mounting_location'][0].upper(),
                unique_report_name + ".mcdx",
            ]
        except Exception as e:
            debug.log(f"save_eqpt_to_csv threw exception: {e}")
        csv_writer.writerow(new_row) #insert the new row
        debug.log(f"save_eqpt_to_csv wrote row: {new_row}")
    f.close() #necissary to close the file
    return True


def get_eqpt_from_xl(filepath: str) -> Equipment:
    """
    Gets all the equipment from an excel file and returns an equipment object
    """
    debug = DebugLogger() 
    wb = load_workbook(filename=filepath)
    try: sheet = wb['values'] # set worksheet to the values sheet
    except: sheet = wb['Calculation']
    # iterate through each of the equipment and append it to the object
    headers = list()
    equipment = Equipment() # new equipment class

    """
    Search for header row before finding data
    """
    header_row = None 
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        #gow through each row in the excel worksheet
        if row[0] == "eqpt_name":
            headers = list(row)  # will search for header row, then start to take data from the rest of the rows
            # Clean up headers variable by removing all errananeous items
            while "" in headers:
                headers.remove("")
            while None in headers:
                headers.remove(None)
            while " " in headers:
                headers.remove(" ")
            required_headers = ["eqpt_name", "project_number", 'eqpt_tags', "tags", "eqpt_number", "mounting_location"]
            for required_header in required_headers:
                if required_header not in headers: #error
                    alert = Popup("Error", "You must include the following column headers in your excel input file: eqpt_name, project_number, 'eqpt_tags', tags, eqpt_number, mounting_location")
                    alert.alert() #alert the user of their mistake 
                    return equipment #return a blank equipment 

            header_row = idx  #we found the start of the useful information
        #if we found the header row and the current row is not blank
        elif header_row and row[0] is not None:
            cur_eqpt = dict()
            for i, header in enumerate(headers):
                # populates each eqpt with the input fields and [value, units]
                try:
                    units = header.split("(")[1]
                    units = units.strip(")")
                    field = header.split("(")[0]
                    field = field.strip(" ")
                    comment = sheet.cell(row = header_row+1, column = i+1).comment 
                    from main_build.dependencies import helpers 
                    try: comment = helpers.format_comment(comment.text)
                    except: pass 
                    cur_eqpt[field] = [row[i], units, comment]  # assign cur_eqpt
                except:
                    cur_eqpt[header] = [row[i], "", ""]  # <- blank units
            equipment.append(cur_eqpt)
        elif header_row and row[0] is None:
            break  # want to break out and not view all rows when reached end of eqpt list
        debug.log(f"Loaded equipment from excel: {equipment.items}")
    return equipment

def save_eqpt_to_xl(equipment: Equipment, filepath:str)->bool:
    """
    Saves the equipment back to the excel file
    """
    debug = DebugLogger()
    wb = load_workbook(filename=filepath)
    try: sheet = wb['values'] # set worksheet to the values sheet
    except: sheet = wb['Calculation']
    # iterate through each of the equipment and append it to the object
    headers = list()

    """
    Search for header row before finding data
    """
    header_row_location = None 
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        #gow through each row in the excel worksheet
        if row[0] == "eqpt_name":
            #start saving the equipment 
            header_row_location = idx  
            for idx, item in enumerate(equipment.items): # each equipment 
                col_number = 1 
                for key, val in item.items():
                    """
                    Equipment {'eqpt_name':['asdf', ' '], '...'}
                    """
                    sheet.cell(row = header_row_location + 1 + idx + 1,
                                column = col_number,
                                value = val[0])
                    col_number +=1
            break  
    # excel file might be open - which will cause fall into except
    try:
        wb.save(filepath) 
        popup = Popup("File Saved", f"The inputs were saved successfuly to {filepath}")
        popup.alert()
        debug.log(f"Saved Equipment to excel file: {equipment.items}")
    except:
        popup = Popup("Error", f"Please close {filepath} and save again.") 
        popup.alert()

    
class DebugLogger():
    """
    Allows us to keep track of the user actions and in case of a bug or crash 
    -> User has the ability to send us parts or the whole .debug file
    """
    def __init__(self):
        self.log_file = None # path to the log file 
        self.instantiate_log_file()

    def instantiate_log_file(self)->bool:
        """
        Creates the log file and hides it from the windows file manager"""
        import os
        self.log_file = os.getcwd() + "/.debug"  
        file = open(self.log_file, 'a') # need to open as append so it will create the file if it does not exist already
        os.system('attrib +H *.debug /S') # hides the file from the windows file manager (invisible to user) 
        file.close()

    def read(self)->list:
        """
        Read from the debug file
        """
        try:
            file = open(self.log_file, "r")
            contents = file.read()
            contents = contents.split("\n")
            return contents[0:-1] #need to get rid of the last blank line 
        except:
            return []
    def log(self, to_log:str)->bool:
        """
        Logs the argument to the hidden file 
        Returns a confirmation boolean 
        """
        import os 
        try:
            if os.path.getsize(self.log_file) > 50000: # if larger than 50 kb then cut the file in half (discard old data) 
                self.cut_file_in_half()
            file = open(self.log_file, 'a') # open in appending mode 
            from datetime import datetime
            file.write(f'{datetime.now()}: {to_log}\n') # timestamp and line 
            file.close()
            return True 
        except:
            return False 
            
    def cut_file_in_half(self)->bool:
        """
        Cuts the file in half -- gets called when file is too big
        """
        import os 
        filedata = self.read()
        filedata = filedata[len(filedata)//2:] # cut the array that is the read file in half 
        os.system('attrib -H *.debug /S') # make the file visible so we can write to it 
        f = open(self.log_file, "w")
        f.write("\n".join(filedata) + "\n") # writing to the file 
        f.close()
        os.system('attrib +H *.debug /S') # make it hidden again from the filesystem 

    def clear(self)->bool:
        """
        Clears all debug logs
        """
        import os 
        try:
            os.remove(self.log_file) # delete 
            self.instantiate_log_file() # make new .debug file
            return True 
        except:
            return False 

    def render(self):
        """
        Renders a GUI with multiline text so user can view most recent logs
        """
        import PySimpleGUI as sg 
        debug_arr = self.read()
        debug_arr.reverse()
        debug_info = "\n\n".join(debug_arr)
        sg.theme('Reddit')
        window = sg.Window("Debug Log",[
            [
            sg.Multiline(
                        default_text = debug_info,  
                        s=(120,40), 
                        disabled= True, 
                        autoscroll=True, 
                        key ="ml",
                        enable_events=True
                        )
            ], 
            [sg.Button("Scroll to Bottom", key = "down"), 
            sg.Button("Scroll to Top", key = "up"), 
            sg.Button("Share", 
                        key = "Share",
                        tooltip="Draft an email to the developer with your debug file attached.")]
        ])
        while True: #logic loop
            event, values = window.read()
            if event == 'OK' or event == sg.WIN_CLOSED:
                window.close()
                break 
            else: 
                if event == "down": # scroll to the bottom of the viewer 
                    window.Element('ml').set_vscroll_position(1.0) 
                if event == "up": # scroll to the top of the viewer 
                    window.Element('ml').set_vscroll_position(0) 
                if event == "Share": # send the .debug file in an email  
                    self.export()

    def export(self):
        """
        Exports the debug log
        """
        import win32com.client as win32   
        try:
            outlook = win32.Dispatch('outlook.application') # open outlook 
            mail = outlook.CreateItem(0) # create new mail 
            mail.Attachments.Add(self.log_file) # attach document 
            mail.HTMLBody = "Type your message here. Send this email with the current attachement to Parth Korde PKorde@ThorntonTomasetti.com, Richard Kuo RKuo@ThorntonTomasetti.com, or Theresa Curtis TCurtis@ThorntonTomasetti.com for more help."
            mail.Display(True) # so the drafted mail appears on the screen 
        except Exception as e: 
            self.log(f"Error exporting debug log: {e}")

if __name__ == "__main__":
    d = DebugLogger()
    d.log("Hello world!")
    d.render()
    

