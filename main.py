import os, sys
sys.path.insert(1, os.getcwd() + "\\build\\MathCadPy") #allows inport of mathcad module 
from mathcadpy import Mathcad, Worksheet #loading custom 
# from MathcadPy import Mathcad, Worksheet
import PySimpleGUI as sg
from pathlib import Path, PurePath
from openpyxl import Workbook as xlwkbk
from openpyxl import load_workbook 
from time import sleep 

class Popup():
    def __init__(self, title, message):
        self.title = title
        self.message = message
    def confirm(self)->bool:
        popup = sg.Window(self.title, [
                        [sg.Text("")],
                        [sg.Text(self.message), sg.Button("YES", key = "YES", button_color='green'), sg.Button("NO", key = "NO", button_color='red')],
                        [sg.Text("")],
                        ])
        while True:
            event, values = popup.read()
            if event == "OK" or event == "NO" or event == sg.WIN_CLOSED:
                popup.close()
                return False 
            else:
                popup.close()
                return True 
 

def load_gui():
    sg.theme('Reddit')
    default_input_size = (14, 1)
    layout = [

        [sg.Column([
            [sg.Frame("Choose Excel File", [[sg.FileBrowse(key = "excel_file", enable_events = True), sg.InputText(key = "excel_name", size = (50,1), background_color = 'white', enable_events = True)], ])],
        ]),
        sg.Column([
            [sg.Frame("Choose template", [[sg.FileBrowse(key = "template_file", enable_events = True), sg.InputText(key = "template_name", size = (50,1), background_color = 'white', enable_events = True)], ])],
        ]),
        ],

                  
        [sg.Text("Equipment Name: "), sg.InputText(key = "eqpt_name", size = (30, 1), background_color = "yellow"), sg.Text("Mounting Location: "), sg.Combo(["WALL","CEILING", "FLOOR", "FLOOR+WALL"], enable_events = True, key = "mounting_location")],

        [
        sg.Text("W_p :="), sg.InputText("", size=(4,1), key = "w_p_input"), sg.Text("lbf", size = (60,1)),
        sg.Button("Calculate", key = "calculate"), 
        sg.Text("File Name"), sg.InputText(key = 'save_file_name', size = (15,1)), sg.Text(".mcdx"),
        sg.Frame("", [[
            sg.Button("Generate Report", key = "generate_report", size = (17,1)), 
        ],
        [
            sg.Button("Previous", key="previous", size = (8,1)), sg.Button("Next", key = "next", size = (7,1))
        ],
        [
            sg.Button("Generate Report For All", key = "generate_report_for_all")
        ],
        ])],

        [sg.Column([
            [sg.Frame("SEISMIC PARAMETERS & GEOMETRY", [

                [sg.Text("S_DS:=", size = (6,1)), sg.InputText(size = default_input_size, key = "s_ds_input")],
                [sg.Text("a_p:=", size = (6,1)), sg.InputText(size = default_input_size, key = "a_p_input")],
                [sg.Text("R_p:=", size = (6,1)), sg.InputText(size = default_input_size, key = "r_p_input")],
                [sg.Text("I_p:=", size = (6,1)), sg.InputText(size = default_input_size, key = "i_p_input")],
                [sg.Text("z:=", size = (6,1)), sg.InputText(size = default_input_size, key = "z_input")],
                [sg.Text("h:=", size = (6,1)), sg.InputText(size = default_input_size, key = "h_input")],
                [sg.Text("z/h:=", size = (6,1)), sg.InputText(size = default_input_size, key = "z_h_output", background_color='yellow', text_color='black')],
                [sg.Text("")],
                [
                    sg.Text("A:=", size = (4,1)), sg.InputText(size = default_input_size, key = "capital_a_input"), sg.Text("in", size = (6,1)), 
                    sg.Text("B:=", size = (4,1)), sg.InputText(size = default_input_size, key = "capital_b_input"), sg.Text("in", size = (6,1))
                ],
                [
                    sg.Text("a:=", size = (4,1)), sg.InputText(size = default_input_size, key = "a_input"), sg.Text("in", size = (6,1)), 
                    sg.Text("b:=", size = (4,1)), sg.InputText(size = default_input_size, key = "b_input"), sg.Text("in", size = (6,1))
                ],
                [sg.Text("H:=", size = (4, 1)), sg.InputText(size = default_input_size, key = "capital_h_input"), sg.Text("in")],
                [sg.Text("")],
                [sg.Text("CG_y:= 2/3 * H =", size = (15, 1)), sg.InputText(size = default_input_size, key = "cg_y_output", background_color='yellow')], 
                [sg.Text("CG_X1:= A/2 =", size = (15, 1)), sg.InputText(size = default_input_size, key = "cg_x1_output", background_color='yellow')], 
                [sg.Text("CG_X2:= B/2 =", size = (15, 1)), sg.InputText(size = default_input_size, key = "cg_x2_output", background_color='yellow')], 


            ]
            , size = (100,-1))]
            ]),


            #preview images
        sg.Column([
                [sg.Frame("Preview Images", [
                    [sg.Text("Input images")], 
                ])]
            ]),

            ],

        [sg.Column([[
            sg.Frame("DETERMINE SEISMIC FORCE", [ #first column
                [sg.Text("F_p"), 
                    sg.Image(filename = (os.getcwd() + "\\build\images\\rsz_f_p_equation.png")), 
                    sg.Text("="), sg.InputText(size = default_input_size, background_color='yellow', key = "f_p_output"),
                    ],
                [sg.Text("F_p_max"), 
                    sg.Image(filename = (os.getcwd() + "\\build\images\\rsz_f_p_max_equation.png")), 
                    sg.Text("="), sg.InputText(size = default_input_size, background_color='yellow', key = "f_p_max_output"),
                    ],
                [sg.Text("F_p_min"), 
                    sg.Image(filename = (os.getcwd() + "\\build\images\\rsz_f_p_min_equation.png")), 
                    sg.Text("="), sg.InputText(size = default_input_size, background_color='yellow', key = "f_p_min_output"),
                    ],
                [sg.Text("F_p"), 
                    sg.Image(filename = (os.getcwd() + "\\build\images\\rsz_f_p_tot_equation.png")), 
                    sg.Text("="), sg.InputText(size = default_input_size, background_color='yellow', key = "f_p_tot_output"),
                    ],
            
            ]
            ,size = (100, -1))
            ]]), sg.Column([ #second column
               [sg.Frame("DETERMINE SEISMIC FORCE", [
                    [sg.Text("Connection to equipment"), sg.Combo(["ASD", "LRFD"], enable_events = True, key = "connection_to_eqpt")],
                    [sg.Text("Load Combinations:")], 
                    [sg.Text("F_PV:="), sg.InputText("", key = "f_pv_connection_to_eqpt", size = default_input_size, background_color = "yellow"), sg.Text("lbf")],
                    [sg.Text("F_PH:="), sg.InputText("", key = "f_ph_connection_to_eqpt", size = default_input_size, background_color = "yellow"), sg.Text("lbf")],

                   ])], 

            ]), sg.Column([#third column

                [sg.Frame("DETERMINE SEISMIC FORCE", [
                    [sg.Text("Connection to base"), sg.Combo(["ASD", "LRFD"], enable_events = True, key = "connection_to_base")],
                    [sg.Text("Load Combinations:")], 
                    [sg.Text("F_PV:="), sg.InputText("", key = "f_pv_connection_to_base", size = default_input_size, background_color = "yellow"), sg.Text("lbf")],
                    [sg.Text("F_PH:="), sg.InputText("", key = "f_ph_connection_to_base", size = default_input_size, background_color = "yellow"), sg.Text("lbf")],

                    ])]
                ])],

            [sg.Text("Determine Maximum Tension and Shear at Anchor Points")],
            [sg.Column([
                [sg.Text("Connection to Equiptment", size = (30, 1))], 
                [sg.Text("T_max =", key = "eqpt_t_max_det_label"), sg.InputText("", size = default_input_size, key = "eqpt_t_max_u_determine")],
                [sg.Text("V_max =", key = "eqpt_v_max_det_label"), sg.InputText("", size = default_input_size, key = "eqpt_v_max_u_determine")],
            ]),
            sg.Column([
                [sg.Text("Connection to Base")], 
                [sg.Text("T_max =", key = "base_t_max_det_label"), sg.InputText("", size = default_input_size, key = "base_t_max_u_determine")],
                [sg.Text("V_max =", key = "base_v_max_det_label"), sg.InputText("", size = default_input_size, key = "base_v_max_u_determine")],
            ]),
            ],

        [sg.Text("Status: ", background_color = "yellow"), sg.Text("OK", key = "cur_status", size = (20,1), background_color = "yellow")],
    ]

    window = sg.Window('Anchorage Mathcad Automation', layout)
    """Some important values"""
    eqpt_num = 1
    max_rows = 0
    """GUI LOOP"""
    while True: 
        event, values = window.read()
        if event == "OK" or event == sg.WIN_CLOSED:
            break #ends gui 
        else:
            """
            Load an excel file 
            Update the current eqpt being viewed 
            """
            if event == "excel_name" or event == "next" or event == "previous":
                if event == "next":
                    eqpt_num = (eqpt_num + 1) % (max_rows+1)
                    if eqpt_num == 0: eqpt_num = 1
                if event == "previous":
                    eqpt_num = (eqpt_num - 1) % (max_rows+1)
                    if eqpt_num == 0: eqpt_num = max_rows  
                if values["excel_name"] != "":
                    inputs, num_rows = set_inputs_from_xl(values['excel_name'], eqpt_num)
                    max_rows = num_rows -1 
                    for key,val in inputs.items():
                        values[key] = val
                        window[key].update(val)
                window['cur_status'].update(f'Equipment {eqpt_num}/{max_rows} loaded')
                

            """
            Update the maximum tension and shear anchor points labels
            """
            if event == "connection_to_base" or event == "connection_to_eqpt":
                if values["connection_to_base"] == "ASD":
                    window['base_t_max_det_label'].update("T_max=")
                    window['base_v_max_det_label'].update("T_max=")
                if values["connection_to_eqpt"] == "ASD":
                    window['eqpt_t_max_det_label'].update("T_max=")
                    window['eqpt_v_max_det_label'].update("V_max=")
                if values["connection_to_base"] == "LRFD":
                    window['base_t_max_det_label'].update("T_u=")
                    window['base_v_max_det_label'].update("V_u=")
                if values["connection_to_eqpt"] == "LRFD":
                    window['eqpt_t_max_det_label'].update("T_u=")
                    window['eqpt_v_max_det_label'].update("V_u=")


            """
            Generate report from the values input
            """
            if event == "generate_report": #perform calculations  
                if values['excel_name'] == "" or values['template_file'] == "":
                    values['cur_status'] = "Please select files"
                    window['cur_status'].update("Please select files")
                else:
                    values['cur_status'] = "Generating Report"
                    window['cur_status'].update("Generating Report")
                    status = generate_report(values, debug = False)
                    if status:
                        status = "File saved."
                    else:
                        status = "Error saving file."
                    values['cur_status'] = status 
                    window['cur_status'].update(status)

            """
            Generate Report for all eqpt
            """
            if event == "generate_report_for_all":
                if values['excel_name'] == "" or values['template_file'] == "":
                    values['cur_status'] = "Please select files"
                    window['cur_status'].update("Please select files")
                else:
                    values['cur_status'] = "Generating Reports"
                    window['cur_status'].update("Generating Reports")
                    for i in range(1, max_rows+1):
                        print(f'{i}/{max_rows}')
                        inputs, num_rows = set_inputs_from_xl(values['excel_name'], i)
                        max_rows = num_rows -1 
                        for key,val in inputs.items():
                            values[key] = val
                            window[key].update(val)
                        new_name = values['eqpt_name']
                        new_name = new_name.replace(" ", "_")
                        new_name += "_report"
                        values['save_file_name'] = new_name
                        status = generate_report(values, debug = False)
                        if status:
                            status = "File saved."
                        else:
                            status = "Error saving file."
                        values['cur_status'] = status 
                        window['cur_status'].update(status)


            """
            Get the outputs from the mathcad file
            """
            if event == "calculate":
                if values['excel_name'] == "" or values['template_file'] == "":
                    values['cur_status'] = "Please select files"
                    window['cur_status'].update("Please select files")
                else:
                    out = mathcad_calculate(values)
                    for key, val in out.items():
                        #cleanup
                        val = str(val[0]) + str(val[1])
                        val = val.replace("{", "")
                        val = val.replace("}", "")
                        #save
                        values[key] = val
                        window[key].update(val)
            






    window.close(); del window
    return 

def set_inputs_from_xl(filepath: str, eqpt_num:int):
    """
    Gets all the inputs from the excel sheet and returns the proper values packet back 
    """
    inputs = dict()
    wb = load_workbook(filename = filepath)
    sheet = wb['values']
    for i in range(eqpt_num):
        num = str(eqpt_num + 1)
        inputs['eqpt_name'] = sheet['A'+num].value
        inputs['mounting_location'] = sheet['B'+num].value.upper()
        inputs['w_p_input'] = sheet['C' + num].value
        inputs['s_ds_input'] = sheet['D' + num].value
        inputs['a_p_input'] = sheet['E' + num].value
        inputs['r_p_input'] = sheet['F' + num].value
        inputs['i_p_input'] = sheet['G' + num].value
        inputs['z_input'] = sheet['H' + num].value
        inputs['h_input'] = sheet['I' + num].value
        inputs['capital_a_input'] = sheet['J' + num].value
        inputs['capital_b_input'] = sheet['K' + num].value
        inputs['a_input'] = sheet['L' + num].value
        inputs['b_input'] = sheet['M' + num].value
        inputs['capital_h_input'] = sheet['N' + num].value



    
    return inputs, sheet.max_row
     

def mathcad_calculate(values:dict, debug = False)->dict:
    """
    Gets all the inputs and performs calculations, 
    returns a dictionary with the output values 
    """
    new_filepath = values['template_file'].split("/")[0:-1] 
    new_filepath = "/".join(new_filepath)
    new_filepath = new_filepath + "/" + values['save_file_name'] + ".mcdx" 
    mathcad_app = Mathcad(visible = debug)
    cur_worksheet = mathcad_app.open(values['template_file']) 

    tosave = dict()
    inputs = [
            'w_p_input', 's_ds_input',
            'a_p_input', 'r_p_input',
            'i_p_input', 'z_input',
            'h_input', 'capital_a_input',
            'capital_b_input', 'a_input',
            'b_input', 'capital_h_input',
    ]
    for i in inputs:
        tosave[i] = values[i]
    for key, value in tosave.items():
        cur_worksheet.set_real_input(str(key), float(value))
    cur_worksheet.synchronize()
    toout = dict()
    outputs = [
            'z_h_output', 'cg_y_output', 
            'cg_x1_output', 'cg_x2_output',
            'cg_x2_output', 'f_p_output',
            'f_p_max_output', 'f_p_min_output',
            'f_p_tot_output', 

    ]
    for o in outputs:
        toout[o] = cur_worksheet.get_real_output(o)
    if debug == False:
        cur_worksheet.close(2) #closes the worksheet and doesn't save it
    return toout


def set_mathcad_inputs(inputs:dict, worksheet:Worksheet)->bool:
    """
    Gets all the inputs as a dictionary 
    returns True if successful, False if unsuccessful
    """
    try:
        for key in inputs:
            worksheet.set_real_input(str(key), inputs[key])
        return True 
    except:
        return False


def generate_report(values, debug = False)->bool:
    """
    Generates report with input values found in the gui 
    """
    new_filepath = values['template_file'].split("/")[0:-1] 
    new_filepath = "/".join(new_filepath)
    new_filepath = new_filepath + "/" + values['save_file_name'] + ".mcdx" 
    mathcad_app = Mathcad(visible = debug)
    cur_worksheet = mathcad_app.open(values['template_file']) 

    tosave = dict()
    inputs = [
            'w_p_input', 's_ds_input',
            'a_p_input', 'r_p_input',
            'i_p_input', 'z_input',
            'h_input', 'capital_a_input',
            'capital_b_input', 'a_input',
            'b_input', 'capital_h_input',
    ]
    for i in inputs:
        tosave[i] = values[i]
    for key, value in tosave.items():
        try:
            cur_worksheet.set_real_input(str(key), float(value))
        except:
            pass

    if cur_worksheet.save_as(new_filepath):
        cur_worksheet.close()
        return True 
    else:
        cur_worksheet.close()
        return False


if __name__ == "__main__":
    #generate_file(debug = True)
    load_gui()
